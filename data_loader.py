from pathlib import Path

import pandas as pd
from openpyxl import load_workbook as openpyxl_load_workbook
from openpyxl.utils.cell import column_index_from_string

SUPPORTED_EXTENSIONS = {".csv", ".xlsx"}
DATE_HINTS = ("date", "data", "month", "mes", "ano")
ID_NAMES = {"id", "codigo", "code", "cnpj", "cpf"}


def load_workbook(filepath: str) -> dict:
    """Load a CSV or XLSX file into a workbook-like structure."""
    path = Path(filepath)
    if not path.exists():
        raise FileNotFoundError(f"File not found: {filepath}")

    extension = path.suffix.lower()
    if extension not in SUPPORTED_EXTENSIONS:
        supported = ", ".join(sorted(SUPPORTED_EXTENSIONS))
        raise ValueError(f"Unsupported file format: {extension}. Supported: {supported}")

    if extension == ".csv":
        return _load_csv_workbook(path)
    return _load_xlsx_workbook(path)


def load_csv(filepath: str) -> pd.DataFrame:
    """Backward-compatible helper that returns the first sheet for CSV files."""
    workbook = load_workbook(filepath)
    return workbook["sheets"][0]["dataframe"]


def _load_csv_workbook(path: Path) -> dict:
    """Load a single CSV as a workbook with one virtual sheet."""
    dataframe = pd.read_csv(path)
    sheet_name = path.stem or "Sheet1"
    sheet = _build_sheet_record(sheet_name, dataframe, formula_examples=[])
    return {
        "filepath": str(path),
        "filetype": "csv",
        "sheet_count": 1,
        "sheets": [sheet],
    }


def _load_xlsx_workbook(path: Path) -> dict:
    """Load an XLSX workbook and extract values plus formula metadata."""
    excel_file = pd.ExcelFile(path)
    formula_map = _extract_formula_map(path)
    sheets = []
    for sheet_name in excel_file.sheet_names:
        dataframe = excel_file.parse(sheet_name=sheet_name)
        formulas = formula_map.get(sheet_name, [])
        sheets.append(_build_sheet_record(sheet_name, dataframe, formulas))

    return {
        "filepath": str(path),
        "filetype": "xlsx",
        "sheet_count": len(sheets),
        "sheets": sheets,
    }


def _extract_formula_map(path: Path) -> dict[str, list[dict]]:
    """Collect formula counts and examples per sheet using openpyxl."""
    workbook = openpyxl_load_workbook(path, data_only=False, read_only=True)
    formula_map: dict[str, list[dict]] = {}
    try:
        for worksheet in workbook.worksheets:
            formulas = []
            for row in worksheet.iter_rows():
                for cell in row:
                    if cell.data_type == "f":
                        formulas.append({
                            "cell": cell.coordinate,
                            "formula": str(cell.value),
                        })
            formula_map[worksheet.title] = formulas
    finally:
        workbook.close()
    return formula_map


def _build_sheet_record(sheet_name: str, dataframe: pd.DataFrame, formula_examples: list[dict]) -> dict:
    """Build a normalized record for a single sheet."""
    normalized = _normalize_dataframe(dataframe)
    formula_columns = _extract_formula_columns(normalized["dataframe"], formula_examples)
    summary = get_column_summary(normalized["dataframe"], formula_columns=formula_columns)
    preview = normalized["dataframe"][summary["usable_columns"][:6]].head(3).fillna("").to_dict("records")
    return {
        "name": sheet_name,
        "dataframe": normalized["dataframe"],
        "raw_shape": dataframe.shape,
        "shape": normalized["dataframe"].shape,
        "empty_rows_removed": normalized["empty_rows_removed"],
        "column_summary": summary,
        "formula_count": len(formula_examples),
        "formula_columns": formula_columns,
        "formula_examples": formula_examples[:5],
        "preview": preview,
    }


def _normalize_dataframe(dataframe: pd.DataFrame) -> dict:
    """Clean fully empty rows and normalize column names."""
    cleaned = dataframe.copy()
    cleaned.columns = [_normalize_column_name(column, index) for index, column in enumerate(cleaned.columns)]

    if cleaned.empty:
        return {
            "dataframe": cleaned,
            "empty_rows_removed": 0,
        }

    empty_row_mask = cleaned.isna().all(axis=1)
    empty_rows_removed = int(empty_row_mask.sum())
    cleaned = cleaned.loc[~empty_row_mask].reset_index(drop=True)
    cleaned = _maybe_parse_date_columns(cleaned)
    return {
        "dataframe": cleaned,
        "empty_rows_removed": empty_rows_removed,
    }


def _normalize_column_name(column, index: int) -> str:
    """Create stable display names for columns."""
    text = str(column).strip()
    if not text or text.startswith("Unnamed:"):
        return f"column_{index + 1}"
    return text


def _maybe_parse_date_columns(dataframe: pd.DataFrame) -> pd.DataFrame:
    """Parse date-like object columns when the signal is strong enough."""
    for column in dataframe.columns:
        if not pd.api.types.is_object_dtype(dataframe[column]):
            continue

        lowered = column.lower()
        if not any(token in lowered for token in DATE_HINTS):
            continue

        converted = pd.to_datetime(dataframe[column], errors="coerce", format="mixed")
        non_null = dataframe[column].notna().sum()
        if non_null == 0:
            continue
        success_ratio = converted.notna().sum() / non_null
        if success_ratio >= 0.8:
            dataframe[column] = converted
    return dataframe


def _extract_formula_columns(dataframe: pd.DataFrame, formula_examples: list[dict]) -> list[str]:
    """Map formula cell coordinates back to dataframe column names."""
    columns = dataframe.columns.tolist()
    mapped = []
    for item in formula_examples:
        coordinate = item.get("cell", "")
        letters = "".join(char for char in coordinate if char.isalpha())
        if not letters:
            continue
        index = column_index_from_string(letters) - 1
        if 0 <= index < len(columns):
            mapped.append(columns[index])
    return list(dict.fromkeys(mapped))


def _looks_like_id_name(column: str) -> bool:
    """Return True when a column name strongly suggests an identifier."""
    normalized = str(column).strip().lower().replace("-", "_")
    parts = [part for part in normalized.split("_") if part]
    if normalized.endswith("_id"):
        return True
    if parts and parts[-1] in ID_NAMES:
        return True
    if any(part in {"cnpj", "cpf"} for part in parts):
        return True
    return False


def get_column_summary(dataframe: pd.DataFrame, formula_columns: list[str] | None = None) -> dict:
    """Return metadata about columns in a generic sheet."""
    formula_columns = formula_columns or []
    columns = dataframe.columns.tolist()
    numeric_columns = dataframe.select_dtypes(include="number").columns.tolist()
    date_columns = dataframe.select_dtypes(include=["datetime", "datetimetz"]).columns.tolist()
    text_columns = [column for column in columns if column not in numeric_columns and column not in date_columns]
    empty_columns = [column for column in columns if dataframe[column].isna().all() and column not in formula_columns]
    formula_only_columns = [column for column in formula_columns if column in columns and dataframe[column].isna().all()]
    mostly_empty_columns = [
        column
        for column in columns
        if column not in empty_columns and column not in formula_only_columns and not dataframe.empty and dataframe[column].isna().mean() >= 0.8
    ]
    likely_id_columns = []
    for column in columns:
        if not _looks_like_id_name(column):
            continue
        non_null = dataframe[column].dropna()
        if non_null.empty:
            continue
        uniqueness_ratio = non_null.nunique() / len(non_null)
        if uniqueness_ratio >= 0.85 and len(non_null) >= max(3, int(len(dataframe) * 0.4)):
            likely_id_columns.append(column)

    return {
        "shape": dataframe.shape,
        "columns": columns,
        "usable_columns": [column for column in columns if column not in empty_columns],
        "numeric_columns": numeric_columns,
        "date_columns": date_columns,
        "text_columns": text_columns,
        "empty_columns": empty_columns,
        "formula_columns": formula_columns,
        "formula_only_columns": formula_only_columns,
        "mostly_empty_columns": mostly_empty_columns,
        "likely_id_columns": likely_id_columns[:4],
        "dtypes": dataframe.dtypes.astype(str).to_dict(),
    }
