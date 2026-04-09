from pathlib import Path

from openpyxl import Workbook

from analyzer import build_ai_context, build_visual_snapshot, generate_insights, get_snapshot
from data_loader import load_workbook


def _build_sample_xlsx(path: Path) -> Path:
    workbook = Workbook()
    revenue = workbook.active
    revenue.title = "Revenue"
    revenue.append(["month", "revenue", "cost", "gross_profit"])
    revenue.append(["2024-01", 1000, 400, "=B2-C2"])
    revenue.append(["2024-02", 1400, 560, "=B3-C3"])
    revenue.append(["2024-03", 1620, 630, "=B4-C4"])

    customers = workbook.create_sheet("Customers")
    customers.append(["customer_id", "customer_name", "segment", "health_score"])
    customers.append([1, "AtlasPay", "Fintech", 82])
    customers.append([2, "Nexa Saude", "Health", 74])

    workbook.save(path)
    workbook.close()
    return path


def test_generate_insights_returns_workbook_summary():
    workbook = load_workbook("data/sample.csv")
    insights = generate_insights(workbook)

    assert isinstance(insights, str)
    assert "Arquivo csv" in insights
    assert "Abas com maior volume" in insights


def test_get_snapshot_returns_generic_workbook_counts():
    workbook = load_workbook("data/sample.csv")
    snapshot = get_snapshot(workbook)

    assert snapshot["sheet_count"] == 1
    assert snapshot["total_rows"] > 0
    assert snapshot["primary_sheet"] == "sample"


def test_build_visual_snapshot_returns_cards_and_sheet_rows():
    workbook = load_workbook("data/sample.csv")
    visual = build_visual_snapshot(workbook)

    assert len(visual["cards"]) == 4
    assert len(visual["sheets"]) == 1
    assert len(visual["highlights"]) >= 2


def test_build_ai_context_includes_formula_examples_and_sheet_names(tmp_path):
    workbook = load_workbook(str(_build_sample_xlsx(tmp_path / "sample.xlsx")))
    context = build_ai_context(workbook)

    assert "LEITURA HUMANA DO ARQUIVO" in context
    assert "ABA: Revenue" in context
    assert "ABA: Customers" in context
    assert "leitura provavel" in context
    assert "=B2-C2" in context
    assert "AtlasPay" in context
