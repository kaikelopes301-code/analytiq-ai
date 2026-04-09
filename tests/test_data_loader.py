from pathlib import Path

import pandas as pd
import pytest
from openpyxl import Workbook

from data_loader import get_column_summary, load_workbook

CSV_PATH = "data/sample.csv"


def _build_sample_xlsx(path: Path) -> Path:
    workbook = Workbook()
    revenue = workbook.active
    revenue.title = "Revenue"
    revenue.append(["month", "revenue", "cost", "gross_profit", "blank_col"])
    revenue.append(["2024-01", 1000, 400, "=B2-C2", None])
    revenue.append(["2024-02", 1400, 560, "=B3-C3", None])

    customers = workbook.create_sheet("Customers")
    customers.append(["customer_id", "customer_name", "segment"])
    customers.append([1, "AtlasPay", "Fintech"])
    customers.append([2, "Nexa Saude", "Health"])

    workbook.save(path)
    workbook.close()
    return path


def test_load_workbook_returns_single_sheet_for_csv():
    workbook = load_workbook(CSV_PATH)
    assert workbook["filetype"] == "csv"
    assert workbook["sheet_count"] == 1
    assert workbook["sheets"][0]["shape"][0] > 0


def test_load_workbook_raises_on_missing_file():
    with pytest.raises(FileNotFoundError):
        load_workbook("nonexistent.csv")


def test_load_workbook_supports_xlsx_with_multiple_sheets_and_formulas(tmp_path):
    path = _build_sample_xlsx(tmp_path / "sample.xlsx")
    workbook = load_workbook(str(path))

    assert workbook["filetype"] == "xlsx"
    assert workbook["sheet_count"] == 2
    assert workbook["sheets"][0]["name"] == "Revenue"
    assert workbook["sheets"][0]["formula_count"] == 2
    assert "gross_profit" in workbook["sheets"][0]["formula_columns"]
    assert "gross_profit" in workbook["sheets"][0]["column_summary"]["formula_only_columns"]
    assert workbook["sheets"][1]["name"] == "Customers"


def test_get_column_summary_returns_generic_metadata():
    dataframe = pd.DataFrame({
        "date": pd.to_datetime(["2024-01-01", "2024-02-01"]),
        "revenue": [100, 120],
        "customer": ["AtlasPay", "Nexa Saude"],
        "empty": [None, None],
    })
    summary = get_column_summary(dataframe)

    assert "shape" in summary
    assert "numeric_columns" in summary
    assert "date_columns" in summary
    assert "text_columns" in summary
    assert "empty" in summary["empty_columns"]
